#!/usr/bin/env python3
"""
SEMrush Keyword Cleaner - porting Python di Codice.gs, eseguibile in chat
(Claude Code o claude.ai code execution) senza Google Sheets/Drive.

Modalita':
  clean               -> pulisce/deduplica/raggruppa i CSV e produce un .xlsx
  find-missing-brands -> individua varianti di brand non ancora note
"""

import argparse
import csv
import io
import os
import re
import sys
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

for _stream in (sys.stdout, sys.stderr):
    if hasattr(_stream, "reconfigure"):
        _stream.reconfigure(encoding="utf-8", errors="replace")

COLUMN_ALIASES = {
    "keyword": ["keyword", "parola chiave", "keywords", "kw"],
    "position": ["position", "posizione", "pos"],
    "search_volume": ["search volume", "volume", "volume di ricerca", "searches", "sv"],
    "url": ["url", "landing page", "pagina di destinazione", "search result"],
}

MARKET_NAMES = {
    "it": "Italia", "es": "Spagna", "fr": "Francia", "de": "Germania",
    "uk": "UK", "us": "USA", "pt": "Portogallo", "nl": "Olanda",
    "be": "Belgio", "ch": "Svizzera", "at": "Austria", "pl": "Polonia",
}

HEADER_COLORS = ["1F4E79", "1A5276", "154360", "0E3250", "1B4F72", "21618C"]

FILENAME_PATTERN = re.compile(
    r"^(.+?)[-._]organic[-._]Positions[-._]([a-zA-Z]{2})[-._](\d{8})([^.]*)",
    re.IGNORECASE,
)
TIME_PATTERN = re.compile(r"T(\d{2})_?(\d{2})_?(\d{2})", re.IGNORECASE)
TLD_PATTERN = re.compile(r"\.(com|it|es|fr|de|uk|net|org)")

HEADERS_ROW = ["Brand", "Mercato", "Mercato Code", "Data", "Keyword", "Position", "Search Volume", "URL", "Brand/Not Brand"]
COLUMN_ALIGN = ["left", "left", "center", "center", "left", "center", "center", "left", "center"]

THIN_GREY = Side(style="thin", color="DDDDDD")
FULL_BORDER = Border(left=THIN_GREY, right=THIN_GREY, top=THIN_GREY, bottom=THIN_GREY)


class Logger:
    def __init__(self):
        self.rows = []

    def log(self, message):
        ts = datetime.now().strftime("%H:%M:%S")
        print(f"[{ts}] {message}")
        self.rows.append([ts, message])


def split_list(raw, seps=","):
    if not raw:
        return []
    pattern = "[" + re.escape(seps) + "]"
    return [v.strip() for v in re.split(pattern, raw.lower()) if v.strip()]


def parse_brand_override(raw):
    """Converte 'chiave=NomeCorretto,chiave2=NomeCorretto2' in {chiave: NomeCorretto}."""
    mapping = {}
    if not raw:
        return mapping
    for pair in raw.split(","):
        pair = pair.strip()
        if not pair or "=" not in pair:
            continue
        key, _, value = pair.partition("=")
        key = key.strip().lower()
        value = value.strip()
        if key and value:
            mapping[key] = value
    return mapping


def parse_csv_text(text):
    rows = list(csv.reader(io.StringIO(text)))
    if rows and len(rows[0]) == 1:
        rows = list(csv.reader(io.StringIO(text), delimiter=";"))
    return rows


def detect_columns(headers):
    headers_lower = [h.strip().lower() for h in headers]
    col_map = {}
    missing = []
    for key, aliases in COLUMN_ALIASES.items():
        found_idx = -1
        for alias in aliases:
            if alias in headers_lower:
                found_idx = headers_lower.index(alias)
                break
        if found_idx != -1:
            col_map[key] = found_idx
        else:
            missing.append(key)
    return col_map, missing


def levenshtein(a, b):
    if not a:
        return len(b)
    if not b:
        return len(a)
    matrix = [[0] * (len(a) + 1) for _ in range(len(b) + 1)]
    for i in range(len(b) + 1):
        matrix[i][0] = i
    for j in range(len(a) + 1):
        matrix[0][j] = j
    for i in range(1, len(b) + 1):
        for j in range(1, len(a) + 1):
            if b[i - 1] == a[j - 1]:
                matrix[i][j] = matrix[i - 1][j - 1]
            else:
                matrix[i][j] = min(matrix[i - 1][j - 1] + 1, matrix[i][j - 1] + 1, matrix[i - 1][j] + 1)
    return matrix[len(b)][len(a)]


def find_close_tokens(keywords, known_variants, min_len=4, max_distance=2):
    """Trova token nelle keyword a distanza di Levenshtein 1-2 da una variante nota (typo/misspelling)."""
    seeds = [v for v in known_variants if len(v) >= min_len]
    found = set()
    if not seeds:
        return found
    for kw in keywords:
        if any(v in kw for v in known_variants):
            continue
        for token in kw.split():
            if len(token) < min_len:
                continue
            for seed in seeds:
                distance = levenshtein(token, seed)
                if 1 <= distance <= max_distance:
                    found.add(token)
                    break
    return found


def parse_filename(filename):
    match = FILENAME_PATTERN.match(filename)
    if not match:
        return None
    brand, mercato_code, data_raw, raw_suffix = match.group(1), match.group(2), match.group(3), match.group(4) or ""
    mercato_code = mercato_code.lower()
    time_match = TIME_PATTERN.search(raw_suffix)
    ts_part = (time_match.group(1) + time_match.group(2) + time_match.group(3)) if time_match else ""
    mercato_name = MARKET_NAMES.get(mercato_code, mercato_code.upper())
    data_str = f"{data_raw[0:4]}-{data_raw[4:6]}-{data_raw[6:8]}"
    return {
        "brand": brand,
        "mercato_code": mercato_code,
        "mercato_name": mercato_name,
        "data_str": data_str,
        "ts_part": ts_part,
    }


def list_csv_files(input_dir):
    if not os.path.isdir(input_dir):
        return []
    return sorted(f for f in os.listdir(input_dir) if f.lower().endswith(".csv"))


def read_csv_rows(path):
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            with open(path, "r", encoding=encoding, newline="") as fh:
                return parse_csv_text(fh.read())
        except UnicodeDecodeError:
            continue
    raise UnicodeDecodeError("Impossibile decodificare il file", b"", 0, 1, path)


def cmd_clean(args, log):
    input_dir = args.input_dir
    raggruppamento = args.raggruppamento
    tipo_query = args.tipo_query
    brand_varianti_manuali = split_list(args.brand_varianti)
    brand_nome_override = parse_brand_override(args.brand_nome_override)
    salta_rilevamento_varianti = args.salta_rilevamento_varianti
    url_esclusi = split_list(args.url_esclusi)
    url_confronto = split_list(args.url_confronto)

    log.log("🚀 Avvio elaborazione — " + datetime.now(timezone.utc).astimezone().strftime("%Y-%m-%d %H:%M:%S"))
    log.log(f"📁 Cartella input: {input_dir}")
    log.log(f"⚙ Raggruppamento: {raggruppamento} | Query: {tipo_query}")
    log.log("🏷 Varianti brand fornite manualmente: " + (", ".join(brand_varianti_manuali) if brand_varianti_manuali else "nessuna"))
    if brand_nome_override:
        log.log("✏ Nome brand corretto dall'utente: " + ", ".join(f"{k}→{v}" for k, v in brand_nome_override.items()))
    if salta_rilevamento_varianti:
        log.log("🛑 Rilevamento automatico varianti disattivato (lista varianti fornita dall'utente)")
    log.log("🚫 URL esclusi: " + (", ".join(url_esclusi) if url_esclusi else "nessuno"))
    log.log("🔀 URL confronto: " + (", ".join(url_confronto) if url_confronto else "nessuno"))

    files = list_csv_files(input_dir)
    pending_by_group = {}
    keywords_by_brand = {}
    file_count = 0
    primo_brand_rilevato = "SEMrush"

    for filename in files:
        parsed = parse_filename(filename)
        if not parsed:
            log.log(f"⏭ Ignorato (pattern non riconosciuto): {filename}")
            continue

        file_count += 1
        brand = parsed["brand"]
        mercato_code = parsed["mercato_code"]
        mercato_name = parsed["mercato_name"]
        data_str = parsed["data_str"]
        ts_part = parsed["ts_part"]
        brand_display_raw = TLD_PATTERN.sub("", brand).strip()
        brand_key = brand_display_raw.lower()
        brand_display = brand_nome_override.get(brand_key, brand_display_raw)

        if file_count == 1:
            primo_brand_rilevato = brand_display

        log.log(f"📄 File #{file_count}: {filename}")
        log.log(f"   → Brand: {brand} | Mercato: {mercato_code.upper()} | Data: {data_str} | TS: {ts_part or 'n/a'}")

        csv_data = read_csv_rows(os.path.join(input_dir, filename))
        if len(csv_data) < 2:
            log.log("   ⚠ File vuoto o non parsabile, saltato.")
            continue

        log.log(f"   → Righe CSV lette: {len(csv_data) - 1}")

        headers = csv_data[0]
        col_map, missing = detect_columns(headers)
        if missing:
            log.log(f"   ⚠ Colonne mancanti ({', '.join(missing)}), file saltato.")
            continue

        group_key = (brand, mercato_code, mercato_name)
        if raggruppamento == "per-data":
            group_key = group_key + (data_str, ts_part)
        pending_by_group.setdefault(group_key, [])
        keywords_by_brand.setdefault(brand_key, set())

        righe_aggiunte = 0
        righe_scartate = 0
        max_idx = max(col_map.values())

        for row in csv_data[1:]:
            if len(row) <= max_idx:
                continue

            keyword = row[col_map["keyword"]].strip()
            try:
                position = int(row[col_map["position"]])
            except ValueError:
                righe_scartate += 1
                continue

            raw_volume = re.sub(r"[\s.,]", "", row[col_map["search_volume"]])
            try:
                volume = int(raw_volume)
            except ValueError:
                volume = 0

            url = row[col_map["url"]].strip()
            if not keyword:
                righe_scartate += 1
                continue

            lower_url = url.lower()
            if url_esclusi and any(u in lower_url for u in url_esclusi):
                righe_scartate += 1
                continue

            lower_kw = keyword.lower()
            keywords_by_brand[brand_key].add(lower_kw)

            pending_by_group[group_key].append({
                "brand": brand, "brand_display": brand_display, "brand_key": brand_key,
                "mercato_name": mercato_name, "mercato_code": mercato_code.upper(),
                "data_str": data_str, "ts_part": ts_part, "keyword": keyword, "lower_kw": lower_kw,
                "position": position, "volume": volume, "url": url,
            })
            righe_aggiunte += 1

        log.log(f"   ✅ Righe valide: {righe_aggiunte} | Scartate: {righe_scartate} | Totale gruppo: {len(pending_by_group[group_key])}")

    if file_count == 0:
        log.log("❌ Nessun file CSV valido trovato.")
        return None

    log.log(f"📦 File elaborati: {file_count} | Gruppi creati: {len(pending_by_group)}")
    log.log("🏷 Identificazione automatica varianti/misspelling brand...")

    variants_by_brand = {}
    for brand_key, keywords in keywords_by_brand.items():
        seed = set(brand_varianti_manuali)
        if brand_key:
            seed.add(brand_key)
        override_name = brand_nome_override.get(brand_key)
        if override_name:
            seed.add(override_name.lower())

        if salta_rilevamento_varianti:
            auto_trovate = set()
        else:
            auto_trovate = find_close_tokens(keywords, seed)
        variants_by_brand[brand_key] = seed | auto_trovate
        etichetta = brand_nome_override.get(brand_key, brand_key or "(vuoto)")
        if auto_trovate:
            log.log(f"   → Brand '{etichetta}': varianti auto-rilevate: {', '.join(sorted(auto_trovate))}")
        elif salta_rilevamento_varianti:
            log.log(f"   → Brand '{etichetta}': rilevamento automatico disattivato, uso solo varianti fornite: {', '.join(sorted(seed)) or 'nessuna'}")
        else:
            log.log(f"   → Brand '{etichetta}': nessuna variante aggiuntiva rilevata (uso: {', '.join(sorted(seed)) or 'nessuna'})")

    log.log("🧹 Avvio classificazione Brand/Not Brand e deduplicazione...")

    all_data_global = []
    sheets_data = {}

    for group_key, rows in pending_by_group.items():
        current_brand = group_key[0]
        current_mercato_code = group_key[1].upper()
        current_data_str = group_key[3] if len(group_key) > 3 else ""
        current_ts_part = group_key[4] if len(group_key) > 4 else ""

        variants = variants_by_brand.get(rows[0]["brand_key"], set()) if rows else set()

        classified_rows = []
        scartate_tipo = 0
        for row in rows:
            is_brand_kw = any(v in row["lower_kw"] for v in variants)
            row["brand_flag"] = "Brand" if is_brand_kw else "Not Brand"

            if tipo_query == "brand" and not is_brand_kw:
                scartate_tipo += 1
                continue
            if tipo_query == "not-brand" and is_brand_kw:
                scartate_tipo += 1
                continue
            classified_rows.append(row)

        if scartate_tipo:
            log.log(f"   🔎 Filtro '{tipo_query}' su gruppo [{'|||'.join(group_key)}]: {scartate_tipo} righe escluse")

        rows = sorted(classified_rows, key=lambda r: (r["data_str"], r["position"]))

        seen = set()
        cleaned_rows = []
        for row in rows:
            lower_url = row["url"].lower()
            matched_pattern = ""
            for p in url_confronto:
                if p in lower_url:
                    matched_pattern = p
                    break

            unique_key = (row["keyword"].lower(), row["data_str"], matched_pattern)
            if unique_key not in seen:
                seen.add(unique_key)
                cleaned_rows.append(row)
                all_data_global.append(row)

        brand_short_raw = TLD_PATTERN.sub("", current_brand).strip()
        brand_short = brand_nome_override.get(brand_short_raw.lower(), brand_short_raw)

        if current_data_str:
            ts_suffix = (" " + current_ts_part[:6]) if current_ts_part else ""
            max_brand_len = max(31 - 18 - len(ts_suffix), 1)
            sheet_name = f"{brand_short[:max_brand_len]} - {current_mercato_code} - {current_data_str}{ts_suffix}"
        else:
            sheet_name = f"{brand_short} - {current_mercato_code}"[:31]

        log.log(f"📋 Gruppo [{'|||'.join(group_key)}] → Foglio: \"{sheet_name}\" | KW dopo dedup: {len(cleaned_rows)} (da {len(rows)})")
        sheets_data[sheet_name] = cleaned_rows

    all_data_global.sort(key=lambda r: (r["brand"], r["mercato_code"], r["data_str"], r["position"]))
    log.log(f"🌍 Totale righe globali (dopo dedup): {len(all_data_global)}")

    modalita_suffisso = "_Dettagliato" if raggruppamento == "per-data" else "_Consolidato"
    filtro_suffisso = {"brand": "_SoloBrand", "not-brand": "_NotBrand"}.get(tipo_query, "_Tutte")
    data_oggi = datetime.now().strftime("%Y-%m-%d")
    final_report_name = f"Report_{primo_brand_rilevato}{modalita_suffisso}{filtro_suffisso}_{data_oggi}"
    log.log(f"📂 Nome report: {final_report_name}")

    output_path = args.output or os.path.join("output", final_report_name + ".xlsx")
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    wb = Workbook()
    used_names = set()

    log.log(f"✍ Scrittura tab 'Tutti i Dati' ({len(all_data_global)} righe)...")
    ws = wb.active
    ws.title = "Tutti i Dati"
    used_names.add(ws.title)
    write_sheet(ws, all_data_global, 0)

    color_idx = 1
    total_sheets = len(sheets_data)
    for sheet_name, rows in sheets_data.items():
        final_name = sheet_name
        suffix = 2
        while final_name in used_names:
            final_name = f"{sheet_name[:28]} {suffix}"
            suffix += 1
        used_names.add(final_name)

        log.log(f"✍ ({color_idx}/{total_sheets}) Scrittura foglio: \"{final_name}\" → {len(rows)} righe")
        ws = wb.create_sheet(title=final_name)
        write_sheet(ws, rows, color_idx)
        color_idx += 1

    write_log_sheet(wb, log.rows)

    wb.save(output_path)
    log.log(f"✅ Elaborazione completata! Fogli creati: {total_sheets + 1} + LOG")
    log.log(f"💾 Output salvato in: {output_path}")
    return output_path


def write_sheet(ws, rows, color_idx):
    header_color = HEADER_COLORS[color_idx % len(HEADER_COLORS)]
    header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center")

    num_cols = len(HEADERS_ROW)
    for c, title in enumerate(HEADERS_ROW, start=1):
        cell = ws.cell(row=1, column=c, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = FULL_BORDER
    ws.row_dimensions[1].height = 20

    data_font = Font(name="Arial", size=10)
    light_blue = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
    white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    for i, row in enumerate(rows, start=2):
        values = [row["brand_display"], row["mercato_name"], row["mercato_code"], row["data_str"], row["keyword"],
                  row["position"], row["volume"], row["url"], row["brand_flag"]]
        fill = light_blue if i % 2 == 0 else white
        for c, (value, align) in enumerate(zip(values, COLUMN_ALIGN), start=1):
            cell = ws.cell(row=i, column=c, value=value)
            cell.font = data_font
            cell.fill = fill
            cell.alignment = Alignment(horizontal=align, vertical="center")
            cell.border = FULL_BORDER

    num_rows = len(rows) + 1
    for c in range(1, num_cols + 1):
        max_len = len(str(HEADERS_ROW[c - 1]))
        for row in rows:
            values = [row["brand_display"], row["mercato_name"], row["mercato_code"], row["data_str"], row["keyword"],
                      row["position"], row["volume"], row["url"], row["brand_flag"]]
            max_len = max(max_len, len(str(values[c - 1])))
        width = min(max(max_len + 2, 10), 60)
        ws.column_dimensions[get_column_letter(c)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(num_cols)}{num_rows}"


def write_log_sheet(wb, log_rows):
    ws = wb.create_sheet(title="LOG")
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    for c, title in enumerate(["Timestamp", "Messaggio"], start=1):
        cell = ws.cell(row=1, column=c, value=title)
        cell.fill = header_fill
        cell.font = header_font
    for i, (ts, message) in enumerate(log_rows, start=2):
        ws.cell(row=i, column=1, value=ts)
        ws.cell(row=i, column=2, value=message)
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 90
    ws.freeze_panes = "A2"


def cmd_detect_brand(args, log):
    """Rileva solo i nomi brand dai nomi file (nessuna lettura del contenuto CSV) — usato per far
    confermare/correggere il nome brand all'utente prima di procedere con la pulizia."""
    input_dir = args.input_dir
    files = list_csv_files(input_dir)

    brands = {}
    file_count = 0
    for filename in files:
        parsed = parse_filename(filename)
        if not parsed:
            log.log(f"⏭ Ignorato (pattern non riconosciuto): {filename}")
            continue
        file_count += 1
        brand_display = TLD_PATTERN.sub("", parsed["brand"]).strip()
        brand_key = brand_display.lower()
        info = brands.setdefault(brand_key, {"display": brand_display, "files": []})
        info["files"].append(filename)

    if file_count == 0:
        log.log("❌ Nessun file CSV valido trovato.")
        return {}

    log.log(f"📦 File analizzati: {file_count} | Brand distinti rilevati: {len(brands)}")
    for brand_key, info in sorted(brands.items()):
        log.log(f"🏷 Brand rilevato: '{info['display']}' (chiave: {brand_key}) — {len(info['files'])} file")

    return brands


def cmd_detect_varianti(args, log):
    """Rileva le varianti/misspelling automatiche del brand (confermato/corretto dall'utente al passo
    precedente) leggendo le keyword dei CSV — usato per far confermare/correggere la lista all'utente."""
    input_dir = args.input_dir
    brand_varianti_manuali = split_list(args.brand_varianti)
    brand_nome_override = parse_brand_override(args.brand_nome_override)

    files = list_csv_files(input_dir)
    keywords_by_brand = {}
    file_count = 0

    for filename in files:
        parsed = parse_filename(filename)
        if not parsed:
            continue
        brand_display_raw = TLD_PATTERN.sub("", parsed["brand"]).strip()
        brand_key = brand_display_raw.lower()

        csv_data = read_csv_rows(os.path.join(input_dir, filename))
        if len(csv_data) < 2:
            continue

        headers_lower = [h.strip().lower() for h in csv_data[0]]
        kw_idx = -1
        for alias in COLUMN_ALIASES["keyword"]:
            if alias in headers_lower:
                kw_idx = headers_lower.index(alias)
                break
        if kw_idx == -1:
            log.log(f"⚠ Colonna Keyword non trovata in {filename}, saltato per il rilevamento varianti.")
            continue

        file_count += 1
        keywords_by_brand.setdefault(brand_key, set())
        for row in csv_data[1:]:
            if len(row) <= kw_idx:
                continue
            kw = row[kw_idx].strip().lower()
            if kw:
                keywords_by_brand[brand_key].add(kw)

    if file_count == 0:
        log.log("❌ Nessun file CSV valido trovato.")
        return {}

    result = {}
    for brand_key, keywords in keywords_by_brand.items():
        seed = set(brand_varianti_manuali)
        seed.add(brand_key)
        override_name = brand_nome_override.get(brand_key)
        if override_name:
            seed.add(override_name.lower())

        auto_trovate = find_close_tokens(keywords, seed)
        display = brand_nome_override.get(brand_key, brand_key)
        result[brand_key] = sorted(auto_trovate)
        if auto_trovate:
            log.log(f"🏷 Brand '{display}': varianti auto-rilevate: {', '.join(sorted(auto_trovate))}")
        else:
            log.log(f"🏷 Brand '{display}': nessuna variante aggiuntiva rilevata.")

    return result


def cmd_find_missing_brands(args, log):
    input_dir = args.input_dir
    known_brands = split_list(args.brand_varianti, seps=",|")

    files = list_csv_files(input_dir)
    missing_variants = set()
    file_count = 0

    for filename in files:
        parsed = parse_filename(filename)
        if not parsed:
            continue
        file_count += 1
        file_brand = TLD_PATTERN.sub("", parsed["brand"].lower()).strip()

        log.log(f"🔍 ({file_count}) Analisi file: {file_brand}")

        is_covered = any(kb in file_brand or file_brand in kb for kb in known_brands)
        if not is_covered and file_brand:
            missing_variants.add(file_brand)

        try:
            csv_data = read_csv_rows(os.path.join(input_dir, filename))
        except (UnicodeDecodeError, OSError):
            continue
        if len(csv_data) < 2:
            continue

        headers_lower = [h.strip().lower() for h in csv_data[0]]
        kw_idx = -1
        for alias in COLUMN_ALIASES["keyword"]:
            if alias in headers_lower:
                kw_idx = headers_lower.index(alias)
                break
        if kw_idx == -1:
            continue

        keywords = []
        for row in csv_data[1:301]:
            if len(row) <= kw_idx:
                continue
            kw = row[kw_idx].strip().lower()
            if kw:
                keywords.append(kw)

        missing_variants |= find_close_tokens(keywords, known_brands)

    if file_count == 0:
        log.log("❌ Nessun file CSV valido trovato.")
        return []

    final_missing = sorted(v for v in missing_variants if v not in known_brands)
    if final_missing:
        log.log("🔍 Varianti mancanti trovate: " + ", ".join(final_missing))
    else:
        log.log("🔍 Nessuna variante mancante rilevata.")
    return final_missing


def build_parser():
    parser = argparse.ArgumentParser(description="SEMrush Keyword Cleaner (porting Python di Codice.gs)")
    parser.add_argument("--mode", choices=["clean", "find-missing-brands", "detect-brand", "detect-varianti"], default="clean")
    parser.add_argument("--input-dir", default="input", help="Cartella con i CSV SEMrush da elaborare")
    parser.add_argument("--output", default=None, help="Percorso file .xlsx di output (solo --mode clean)")
    parser.add_argument("--raggruppamento", choices=["consolidato", "per-data"], default="consolidato",
                         help="consolidato = un foglio per brand/mercato; per-data = un foglio per brand/mercato/data")
    parser.add_argument("--tipo-query", choices=["tutte", "brand", "not-brand"], default="tutte")
    parser.add_argument("--brand-varianti", default="",
                         help="Varianti brand aggiuntive separate da virgola (es. 'falconeri,falco neri'). "
                              "Il nome brand rilevato dal nome file e le sue varianti/misspelling vengono "
                              "sempre identificati e usati automaticamente per la colonna Brand/Not Brand; "
                              "questo parametro serve solo a integrare varianti che l'euristica potrebbe non trovare.")
    parser.add_argument("--brand-nome-override", default="",
                         help="Correzioni al nome brand rilevato dal nome file, formato 'chiave=NomeCorretto' "
                              "separate da virgola (es. 'falconeri=Falconeri'). La chiave e' il nome brand "
                              "rilevato dal filename in minuscolo (vedi --mode detect-brand). Usato per applicare "
                              "la correzione confermata dall'utente al nome brand prima della pulizia definitiva.")
    parser.add_argument("--salta-rilevamento-varianti", action="store_true",
                         help="Disattiva il rilevamento automatico di typo/misspelling: usa solo il nome brand "
                              "(eventualmente corretto) e le varianti fornite via --brand-varianti. Da usare quando "
                              "l'utente ha fornito una lista di varianti corretta in sostituzione di quella auto-rilevata.")
    parser.add_argument("--url-esclusi", default="", help="Sottostringhe URL da escludere, separate da virgola")
    parser.add_argument("--url-confronto", default="", help="Sottostringhe URL per il confronto in fase di dedup, separate da virgola")
    return parser


def main(argv=None):
    parser = build_parser()
    args = parser.parse_args(argv)
    log = Logger()

    if args.mode == "clean":
        output_path = cmd_clean(args, log)
        if output_path is None:
            sys.exit(1)
    elif args.mode == "find-missing-brands":
        cmd_find_missing_brands(args, log)
    elif args.mode == "detect-brand":
        cmd_detect_brand(args, log)
    else:
        cmd_detect_varianti(args, log)


if __name__ == "__main__":
    main()
